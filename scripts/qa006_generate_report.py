from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
QA_DIR = ROOT / "docs" / "qa-reports" / "wiz-qa-006"
CASES_PATH = QA_DIR / "cases.json"
XLSX_PATH = ROOT / "docs" / "qa-reports" / "QA-Test-006.xlsx"
SUMMARY_MD = QA_DIR / "QA-006-SUMMARY.md"

ROW_HEADERS = [
    "Test ID",
    "Scenario",
    "Steps",
    "Expected",
    "Actual",
    "Status",
    "Severity",
    "URL",
    "Assertions",
    "API Evidence",
    "Screenshot",
    "Console Log",
    "Network Log",
    "Probable Cause",
    "Suggested Fix",
    "QA-005 Reference",
]


def style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F766E")


def cell(v: object) -> str:
    if v is None:
        return ""
    s = re.sub(r"\x1b\[[0-9;]*m", "", str(v))
    return "".join(ch for ch in s if ch in "\n\t" or ord(ch) >= 32)


def sheet_for_batch(wb: Workbook, title: str, batch: str, cases: list[dict]) -> None:
    ws = wb.create_sheet(title)
    style_header(ws, ROW_HEADERS)
    for c in cases:
        if c.get("batch") == batch:
            ws.append(
                [
                    cell(c.get("id")),
                    cell(c.get("scenario")),
                    cell(c.get("steps")),
                    cell(c.get("expected")),
                    cell(c.get("actual")),
                    cell(c.get("status")),
                    cell(c.get("severity")),
                    cell(c.get("url")),
                    cell(c.get("assertions")),
                    cell(c.get("apiEvidence")),
                    cell(c.get("screenshot")),
                    cell(c.get("consoleLog")),
                    cell(c.get("networkLog")),
                    cell(c.get("probableCause")),
                    cell(c.get("suggestedFix")),
                    cell(c.get("qa005Ref")),
                ]
            )


def main() -> None:
    if not CASES_PATH.exists():
        raise SystemExit(f"Run QA-006 Playwright batches first ({CASES_PATH} missing).")

    raw: list[dict] = json.loads(CASES_PATH.read_text(encoding="utf-8"))
    # Keep latest result per (batch, scenario) when batches were re-run
    latest: dict[tuple[str, str], dict] = {}
    for c in raw:
        key = (c.get("batch", ""), c.get("scenario", ""))
        prev = latest.get(key)
        if not prev or c.get("id", "") > prev.get("id", ""):
            latest[key] = c
    cases = list(latest.values())
    cases.sort(key=lambda x: x.get("id", ""))
    total = len(cases)
    passed = sum(1 for c in cases if c["status"] == "PASS")
    failed = sum(1 for c in cases if c["status"] == "FAIL")
    blocked = sum(1 for c in cases if c["status"] == "BLOCKED")

    wb = Workbook()
    s1 = wb.active
    s1.title = "Test Summary"
    style_header(s1, ["Metric", "Value"])
    for row in [
        ("QA Cycle", "QA-006 Targeted Retest"),
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Blocked", blocked),
        ("Inbox Retest Pass", sum(1 for c in cases if c["batch"] == "inbox" and c["status"] == "PASS")),
        ("Workflow Retest Pass", sum(1 for c in cases if c["batch"] == "workflow" and c["status"] == "PASS")),
        ("Session Retest Pass", sum(1 for c in cases if c["batch"] == "session" and c["status"] == "PASS")),
        ("Long-Duration Pass", sum(1 for c in cases if c["batch"] == "longdur" and c["status"] == "PASS")),
        ("Smoke Pass", sum(1 for c in cases if c["batch"] == "smoke" and c["status"] == "PASS")),
        ("Evidence Folder", str(QA_DIR)),
        ("Proceed to QA-007", "Yes" if failed == 0 else "After fixing remaining failures"),
    ]:
        s1.append(list(row))

    sheet_for_batch(wb, "Inbox Sync Retest", "inbox", cases)
    sheet_for_batch(wb, "Workflow Sync Retest", "workflow", cases)
    sheet_for_batch(wb, "Session Interruption Retest", "session", cases)
    sheet_for_batch(wb, "Long-Duration Stability Retest", "longdur", cases)
    sheet_for_batch(wb, "Regression Smoke Test", "smoke", cases)

    s7 = wb.create_sheet("Remaining Issues")
    style_header(s7, ["ID", "Batch", "Scenario", "Severity", "Status", "Summary"])
    n = 1
    for c in cases:
        if c["status"] != "PASS":
            s7.append([f"RI-{n:03d}", c["batch"], c["scenario"], c["severity"], c["status"], cell(c["actual"])[:200]])
            n += 1
    if n == 1:
        s7.append(["RI-000", "—", "No remaining failures", "—", "—", "All targeted retests passed"])

    s8 = wb.create_sheet("Fix Recommendations")
    style_header(s8, ["Priority", "Area", "Recommendation", "Status"])
    recs = [
        ("P0", "Inbox sync", "Use data-testid inbox-list-item and data-inbox-count for QA; API must use limit=100", "Implemented"),
        ("P0", "Workflow sync", "Use data-testid workflow-list-item and data-workflow-count", "Implemented"),
        ("P1", "Session", "Network offline banner + online refreshUser listener", "Implemented"),
        ("P1", "Long-duration", "Fresh browser context per segment; count errors per segment only", "Implemented in QA-006 runner"),
        ("P2", "QA-005", "Prior inbox failures were false positives from bad selector (w-full buttons)", "Documented"),
    ]
    for r in recs:
        s8.append(list(r))

    s9 = wb.create_sheet("Evidence Index")
    style_header(s9, ["Evidence ID", "Test ID", "Type", "Path", "Description"])
    eid = 1
    for c in cases:
        if c.get("screenshot"):
            s9.append([f"E-{eid:04d}", c["id"], "screenshot", c["screenshot"], c["scenario"]])
            eid += 1
    for f in sorted((QA_DIR / "logs").glob("*.log")) if (QA_DIR / "logs").exists() else []:
        s9.append([f"E-{eid:04d}", "N/A", "log", str(f), f.name])
        eid += 1

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)

    fixed = [
        "Added inbox/workflow list test IDs and count attributes for reliable sync validation",
        "Added NetworkStatusBanner and online session refresh in AuthContext",
    ]
    remaining = [f"{c['id']}: {c['scenario']} — {c['actual'][:100]}" for c in cases if c["status"] != "PASS"]

    summary = f"""# QA-006 Targeted Retest Summary

- **Total:** {total} | **Pass:** {passed} | **Fail:** {failed} | **Blocked:** {blocked}
- **Excel:** `{XLSX_PATH}`

## Fixed in this cycle
{chr(10).join('- ' + x for x in fixed)}

## Remaining issues
{chr(10).join('- ' + x for x in remaining) if remaining else '- None'}

## QA-007 recommendation
{"**Proceed** to full QA-007 after user verifies localhost." if failed == 0 else "**Hold** full QA-007 until remaining issues are resolved."}
"""
    SUMMARY_MD.write_text(summary, encoding="utf-8")
    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {SUMMARY_MD}")


if __name__ == "__main__":
    main()
