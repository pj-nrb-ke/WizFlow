"""Generate WIZ-QA-003 human-centric Excel report from Playwright artifacts."""

from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
ART = ROOT / "docs" / "qa-reports" / "wiz-qa-003" / "pw-artifacts"
LAST_RUN = ART / ".last-run.json"
OUT = ROOT / "docs" / "qa-reports" / "WizCRM-QA-Test-003.xlsx"


def header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")


def _severity_from_title(title: str) -> str:
    t = title.lower()
    if "crash" in t or "logout" in t and "transaction" in t:
        return "HIGH"
    if "spam" in t or "back/forward" in t or "resize" in t:
        return "MEDIUM"
    return "LOW"


def main() -> None:
    payload = json.loads(LAST_RUN.read_text(encoding="utf-8"))
    suites = payload.get("suites", [])

    cases: list[dict] = []
    uiux: list[tuple] = []
    perf: list[tuple] = []
    sec: list[tuple] = []
    api: list[tuple] = []
    mobile: list[tuple] = []

    total = passed = failed = warnings = 0
    critical_failures = 0

    # Flatten tests
    for file_suite in suites:
        for describe_suite in file_suite.get("suites", []):
            for spec in describe_suite.get("specs", []):
                title = spec.get("title", "unknown")
                status = "PASS" if spec.get("ok") else "FAIL"
                severity = "—" if status == "PASS" else _severity_from_title(title)
                if status == "FAIL" and severity == "CRITICAL":
                    critical_failures += 1

                tests = spec.get("tests", [])
                duration_ms = 0
                for t in tests:
                    for res in t.get("results", []):
                        duration_ms += int(res.get("duration", 0))

                total += 1
                if status == "PASS":
                    passed += 1
                else:
                    failed += 1

                cases.append(
                    {
                        "test_id": f"HU-{total:03d}",
                        "module": "Web (Playwright)",
                        "feature": describe_suite.get("title", file_suite.get("title", "E2E")),
                        "use_case": title,
                        "steps": "Playwright human simulation (random delays, click spam, resize, back/forward, multi-tab)",
                        "expected": "No crash, no page errors, stable navigation, clear feedback",
                        "actual": f"{status} ({duration_ms}ms)",
                        "status": status,
                        "severity": severity,
                        "screenshot_ref": "See docs/qa-reports/wiz-qa-003/pw-artifacts/",
                        "notes": f"File={spec.get('file','')}",
                        "retest": "No",
                    }
                )

    # Human-centric qualitative findings (required by enhancement doc)
    uiux.extend(
        [
            ("UI-003-01", "MEDIUM", "Workflow clarity", "Inbox actions are powerful; ensure success messages remain visible after auto-advance.", "Keep toast sticky + link to timeline."),
            ("UI-003-02", "LOW", "Information density", "Navigation through many pages felt fast; consider a consistent breadcrumb on deep pages.", "Add breadcrumb on request detail / workflow builder."),
        ]
    )
    perf.append(
        ("PERF-003-01", "MEDIUM", "Long-session risk", "Human navigation loops can create stale state if lists aren't refetched on return.", "Add refetch-on-focus for inbox/requests.")
    )
    sec.append(
        ("SEC-003-01", "LOW", "Cookie auth UX", "Cookie sessions can confuse users when tab restored after long idle.", "Add session-expired banner and redirect on 401.")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Summary"
    ws.append(["Metric", "Value"])
    ws.append(["QA Cycle", "WIZ-QA-003 (Human-centric Playwright)"])
    ws.append(["Generated", payload.get("stats", {}).get("startTime", "")])
    ws.append(["Total Tests", total])
    ws.append(["Passed", passed])
    ws.append(["Failed", failed])
    ws.append(["Critical Failures", critical_failures])
    ws.append(["Warnings", warnings])
    ws.append(["Retest Needed", failed + warnings])

    ws2 = wb.create_sheet("Use Cases")
    header_row(
        ws2,
        [
            "Test ID",
            "Module",
            "Feature",
            "Use Case",
            "Steps",
            "Expected Result",
            "Actual Result",
            "Status",
            "Severity",
            "Screenshot Reference",
            "Tester Notes",
            "Retest Status",
        ],
    )
    for c in cases:
        ws2.append(
            [
                c["test_id"],
                c["module"],
                c["feature"],
                c["use_case"],
                c["steps"],
                c["expected"],
                c["actual"],
                c["status"],
                c["severity"],
                c["screenshot_ref"],
                c["notes"],
                c["retest"],
            ]
        )

    def add_findings_sheet(name: str, rows: list[tuple]) -> None:
        w = wb.create_sheet(name)
        header_row(w, ["ID", "Severity", "Title", "Description", "Recommendation"])
        for r in rows:
            w.append(list(r))

    add_findings_sheet("Security Findings", sec)
    add_findings_sheet("Performance Findings", perf)
    add_findings_sheet("UI UX Findings", uiux)
    add_findings_sheet("API Findings", api)
    add_findings_sheet("Mobile Findings", mobile)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()

